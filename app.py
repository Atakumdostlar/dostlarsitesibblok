"""
from flask import Flask, render_template, request, redirect, session, jsonify, make_response, send_file, flash
from flask_login import LoginManager, login_required, current_user
from functools import wraps
from flask import session, redirect, url_for
"""

from flask import Flask, render_template, request, redirect, session, jsonify, make_response, send_file, flash, url_for
from flask_login import LoginManager, login_required, current_user
from functools import wraps

import sqlite3


from datetime import datetime
from io import BytesIO
import pandas as pd
from weasyprint import HTML

app = Flask(__name__)
app.secret_key = "gizli_anahtar"

def db_connect():
    conn = sqlite3.connect("apartman.db")
    conn.row_factory = sqlite3.Row
    print("Row factory def içindeki:", conn.row_factory)
    return conn

def sadece_yonetici(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if session.get('rol') != 'yonetici':
            return redirect(url_for('giris'))  # Yanlış rol → giriş ekranına gönder
        return f(*args, **kwargs)
    return wrapper


def format_bakiye(bakiye):
    return f"{int(round(bakiye)):,}".replace(",", ".")  # 1234567.89 → 1.234.568

def aidat_id_uret(user_id, yil, ay):
    baslangic = "5"  # sistem tanımı (örneğin apartman aidat sistemi)
    uid_str = f"{int(user_id):03d}"     # 15 → "015"
    yil_str = str(yil)                  # 2025
    ay_str = f"{int(ay):02d}"           # 7 → "07"

    aidat_id = int(f"{baslangic}{uid_str}{yil_str}{ay_str}")
    return aidat_id




@app.route('/')
def home():
    return render_template('login.html')



@app.route('/giris', methods=['POST'])
def giris():
    daire_no = request.form['daire_no']
    sifre = request.form['sifre']

    conn = db_connect()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    cur.execute("SELECT * FROM users WHERE daire_no=? AND sifre=?", (daire_no, sifre))
    user = cur.fetchone()
    conn.close()

    if user:
        session['rol'] = user['rol']
        session['user_id'] = user['user_id']
        return redirect('/panel')  # İstersen role göre yönlendirme ekleyebiliriz
    else:
        return render_template("login.html", hata="Daire numarası veya şifre yanlış!")


@app.route('/panel')
def panel():
    if 'rol' not in session:
        return redirect('/')

    if session['rol'] == 'yonetici':
        conn = db_connect()
        cur = conn.cursor()
        cur.execute("SELECT COALESCE(SUM(tutar), 0) FROM aidatlar")
        toplam_aidat = cur.fetchone()[0]
        cur.execute("SELECT COALESCE(SUM(tutar), 0) FROM gelirler")
        toplam_gelir = cur.fetchone()[0]
        cur.execute("SELECT COALESCE(SUM(tutar), 0) FROM giderler")
        toplam_gider = cur.fetchone()[0]
        conn.close()

        kasa = toplam_aidat + toplam_gelir - toplam_gider

        return render_template("panel_yonetici.html",
            toplam_aidat=toplam_aidat,
            toplam_gelir=toplam_gelir,
            toplam_gider=toplam_gider,
            kasa=kasa
        )

    # 🌿 Sakin paneli
    conn = db_connect()
    cur = conn.cursor()

    user_id = session['user_id']
    cur.execute("SELECT ad_soyad, daire_no FROM users WHERE user_id=?", (user_id,))
    kisi = cur.fetchone()

    cur.execute("""
    SELECT tutar, yil, ay, tarih, durum, aciklama
    FROM aidatlar
    WHERE user_id=?
    ORDER BY yil DESC, ay DESC
    """, (user_id,))

    aidatlar = cur.fetchall()

    

    for r in aidatlar:
        print("Yıl:", r["yil"], "Ay:", r["ay"])

    conn.close()


    toplam_odeme = sum(r["tutar"] for r in aidatlar if r["durum"] == "odendi")
    toplam_borc = sum(r["tutar"] for r in aidatlar if r["durum"] != "odendi")

    return render_template("panel_sakin.html",
        kisi=kisi,
        aidatlar=aidatlar,
        toplam_odeme=toplam_odeme,
        toplam_borc=toplam_borc
    )


@app.route('/cikis')
def cikis():
    session.clear()
    return redirect('/')

@app.route('/gelir-ekle', methods=['GET', 'POST'])
@sadece_yonetici
def gelir_ekle():
    if 'rol' not in session or session['rol'] != 'yonetici':
        return redirect('/')
    if request.method == 'POST':
        baslik = request.form['baslik']
        tutar = request.form['tutar']
        tarih = request.form['tarih']
        aciklama = request.form['aciklama']
        conn = db_connect()
        cur = conn.cursor()
        cur.execute("INSERT INTO gelirler (baslik, tutar, tarih, aciklama) VALUES (?, ?, ?, ?)",
                    (baslik, tutar, tarih, aciklama))
        conn.commit()
        conn.close()
        return redirect('/panel')
    return render_template('gelir_ekle.html')


"""
@app.route('/aidat-ekle', methods=['GET', 'POST'])
def aidat_ekle():
    abort(403)

    
    if 'rol' not in session or session['rol'] != 'yonetici':
        return redirect('/')
    conn = db_connect()
    cur = conn.cursor()
    if request.method == 'POST':
        user_id = request.form['user_id']
        tarih = request.form['tarih']
        tutar = request.form['tutar']
        durum = request.form['durum']
        aciklama = request.form.get('aciklama', '')
        cur.execute(
        "INSERT INTO aidatlar (user_id, tarih, tutar, durum, aciklama) VALUES (?, ?, ?, ?, ?)",
        (user_id, tarih, tutar, durum, aciklama)
        )

        conn.commit()
        conn.close()
        return redirect('/panel')
    cur.execute("SELECT user_id, daire_no, ad_soyad FROM users")
    rows = cur.fetchall()
    conn.close()
    kullanicilar = [ { 'id': row[0], 'etiket': f"{row[1]} - {row[2]}" } for row in rows ]
    return render_template("aidat_ekle.html", kullanicilar=kullanicilar)
"""


            
@app.route("/aidatlar")
@sadece_yonetici
def aidat_listele():
    import sqlite3
    from flask import request, render_template, flash
    


    # 🔍 Parametreleri al
    siralama   = request.args.get("siralama", "tarih")
    yon        = request.args.get("yon", "DESC")
    user_id    = request.args.get("user_id")
    tarih1     = request.args.get("tarih1")
    tarih2     = request.args.get("tarih2")
    tek_sakin_var = bool(user_id)
    mail_gonder_var = not tek_sakin_var 

    # 🔧 Sıralama ve yön doğrulama
    allowed_fields = ["tarih", "tutar", "durum", "ad_soyad", "daire_no"]
    if siralama not in allowed_fields:
        siralama = "tarih"
    if yon not in ["ASC", "DESC"]:
        yon = "DESC"

    conn = sqlite3.connect("apartman.db")
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # 🧱 Ana sorgu oluştur
    query = """
        SELECT
            a.aidat_id,
            u.daire_no,
            u.ad_soyad,
            a.tarih,
            a.tutar,
            a.durum,
            a.yil,
            a.ay,
            a.aciklama
        FROM aidatlar a
        JOIN users u ON a.user_id = u.user_id
        WHERE 1=1
    """
    params = []

    # 👤 Kullanıcı filtresi
    if user_id:
        query += " AND u.user_id = ?"
        params.append(user_id)

    # 📆 Tarih aralığı filtresi
    if tarih1 and tarih2:
        query += " AND a.tarih BETWEEN ? AND ?"
        params.extend([tarih1, tarih2])

    # 🔢 Sıralama
    query += f" ORDER BY {siralama} {yon}"

    # 📋 Sorguyu çalıştır
    cur.execute(query, params)
    kayitlar = cur.fetchall()

    # 👥 Tüm sakinleri listele
    cur.execute("SELECT user_id, daire_no, ad_soyad FROM users ORDER BY daire_no")
    sakinler = [{"id": r["user_id"], "etiket": f"{r['daire_no']} - {r['ad_soyad']}"} for r in cur.fetchall()]
    conn.close()

    

    return render_template("aidat_listele.html",
                       kayitlar=kayitlar,
                       sakinler=sakinler,
                       selected_user=int(user_id) if user_id else None,
                       siralama=siralama,
                       yon=yon,
                       tarih1=tarih1,
                       tarih2=tarih2,
                       tek_sakin_var=tek_sakin_var,
                       mail_gonder_var=mail_gonder_var)



@app.route('/gider-ekle', methods=['GET', 'POST'])
@sadece_yonetici
def gider_ekle():
    if 'rol' not in session or session['rol'] != 'yonetici':
        return redirect('/')
    if request.method == 'POST':
        kategori = request.form['kategori']
        tutar = request.form['tutar']
        tarih = request.form['tarih']
        aciklama = request.form.get('aciklama', '')
        conn = db_connect()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO giderler (kategori, tutar, tarih, aciklama)
            VALUES (?, ?, ?, ?)
        """, (kategori, tutar, tarih, aciklama))
        conn.commit()
        conn.close()
        return redirect('/panel')
    return render_template('gider_ekle.html')



@app.route("/export/gelirler")
@sadece_yonetici
def export_gelirler():
    # Filtre ve sıralama parametreleri
    baslangic = request.args.get("baslangic")
    bitis = request.args.get("bitis")
    siralama = request.args.get("siralama", "tarih")
    yon = request.args.get("yon", "DESC")

    # Güvenlik: sadece geçerli alanlar sıralanabilir
    allowed_fields = ["tarih", "tutar", "baslik"]
    if siralama not in allowed_fields:
        siralama = "tarih"
    if yon not in ["ASC", "DESC"]:
        yon = "DESC"

    conn = sqlite3.connect("apartman.db")
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # Sorgu oluştur
    query = "SELECT baslik, tarih, tutar, aciklama FROM gelirler WHERE 1=1"
    params = []

    if baslangic:
        query += " AND tarih >= ?"
        params.append(baslangic)
    if bitis:
        query += " AND tarih <= ?"
        params.append(bitis)

    query += f" ORDER BY {siralama} {yon}"
    cur.execute(query, params)
    gelirler = cur.fetchall()
    conn.close()

    # Excel dosyası oluştur
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Gelir Raporu"

    # Başlık satırı
    ws.append(["Başlık", "Tarih", "Tutar", "Açıklama"])

    # Veri satırları
    for row in gelirler:
        ws.append([row["baslik"], row["tarih"], row["tutar"], row["aciklama"]])

    # Toplam tutar satırı
    toplam = sum([r["tutar"] for r in gelirler]) if gelirler else 0
    ws.append(["", "", toplam, "Toplam Gelir"])

    # Stil önerisi (istersen ekleyebilirim): kalın başlık, TL simgesi, hizalama vs.

    # Yanıt olarak xlsx gönder
    from io import BytesIO
    excel_stream = BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)

    response = make_response(excel_stream.read())
    response.headers["Content-Disposition"] = "attachment; filename=gelir_raporu.xlsx"
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response






@app.route("/export/ozet/pdf")
@sadece_yonetici
def export_kasa_ozet_pdf():
    baslangic = request.args.get("baslangic")
    bitis = request.args.get("bitis")

    conn = sqlite3.connect("apartman.db")
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    gelir_query = "SELECT SUM(tutar) FROM gelirler WHERE 1=1"
    gider_query = "SELECT SUM(tutar) FROM giderler WHERE 1=1"
    gelir_params = []
    gider_params = []

    if baslangic:
        gelir_query += " AND tarih >= ?"
        gider_query += " AND tarih >= ?"
        gelir_params.append(baslangic)
        gider_params.append(baslangic)

    if bitis:
        gelir_query += " AND tarih <= ?"
        gider_query += " AND tarih <= ?"
        gelir_params.append(bitis)
        gider_params.append(bitis)

    cur.execute(gelir_query, gelir_params)
    toplam_gelir = cur.fetchone()[0] or 0

    cur.execute(gider_query, gider_params)
    toplam_gider = cur.fetchone()[0] or 0
    conn.close()

    net_kasa = toplam_gelir - toplam_gider
    def tl_format(m): return "{:,}".format(int(m)).replace(",", ".")

    rendered = render_template("rapor_pdf_ozet.html",
                               toplam_gelir=tl_format(toplam_gelir),
                               toplam_gider=tl_format(toplam_gider),
                               net_kasa=tl_format(net_kasa),
                               baslangic=baslangic,
                               bitis=bitis)

    pdf = HTML(string=rendered).write_pdf()
    response = make_response(pdf)
    response.headers["Content-Disposition"] = "attachment; filename=kasa_ozeti.pdf"
    response.headers["Content-Type"] = "application/pdf"
    return response

@app.route("/export/ozet")
@sadece_yonetici
def export_kasa_ozet():
    baslangic = request.args.get("baslangic")
    bitis = request.args.get("bitis")

    conn = sqlite3.connect("apartman.db")
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # Ayrı ayrı sorgu parametreleri
    gelir_query = "SELECT SUM(tutar) FROM gelirler WHERE 1=1"
    gider_query = "SELECT SUM(tutar) FROM giderler WHERE 1=1"
    gelir_params = []
    gider_params = []

    if baslangic:
        gelir_query += " AND tarih >= ?"
        gider_query += " AND tarih >= ?"
        gelir_params.append(baslangic)
        gider_params.append(baslangic)

    if bitis:
        gelir_query += " AND tarih <= ?"
        gider_query += " AND tarih <= ?"
        gelir_params.append(bitis)
        gider_params.append(bitis)

    # Sorguları çalıştır
    cur.execute(gelir_query, gelir_params)
    toplam_gelir = cur.fetchone()[0] or 0

    cur.execute(gider_query, gider_params)
    toplam_gider = cur.fetchone()[0] or 0
    conn.close()

    net_kasa = toplam_gelir - toplam_gider

    # Excel dosyası oluştur
    from openpyxl import Workbook
    from io import BytesIO
    wb = Workbook()
    ws = wb.active
    ws.title = "Kasa Özeti"

    ws.append(["Başlık", "Tutar (₺)"])
    ws.append(["Toplam Gelir", toplam_gelir])
    ws.append(["Toplam Gider", toplam_gider])
    ws.append(["Net Kasa", net_kasa])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    response = make_response(stream.read())
    response.headers["Content-Disposition"] = "attachment; filename=kasa_ozeti.xlsx"
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response

@app.route('/yillik-ozet')
@sadece_yonetici
def yillik_ozet():
    if 'rol' not in session or session['rol'] != 'yonetici':
        return redirect('/')
    conn = db_connect()
    cur = conn.cursor()
    cur.execute("SELECT strftime('%Y', tarih) AS yil, SUM(tutar) FROM aidatlar GROUP BY yil")
    aidatlar = dict(cur.fetchall())
    cur.execute("SELECT strftime('%Y', tarih) AS yil, SUM(tutar) FROM gelirler GROUP BY yil")
    gelirler = dict(cur.fetchall())
    cur.execute("SELECT strftime('%Y', tarih) AS yil, SUM(tutar) FROM giderler GROUP BY yil")
    giderler = dict(cur.fetchall())
    conn.close()

    yillar = sorted(set(aidatlar) | set(gelirler) | set(giderler))
    ozet = []
    devir = 0

    for yil in yillar:
        aidat = aidatlar.get(yil, 0)
        gelir = gelirler.get(yil, 0)
        gider = giderler.get(yil, 0)
        toplam = aidat + gelir
        kasa = devir + toplam - gider

        ozet.append({
            'yil': yil,
            'aidat': aidat,
            'gelir': gelir,
            'gider': gider,
            'toplam': toplam,
            'devir': devir,
            'kasa': kasa
        })

        devir = kasa

    session["ozet_cache"] = ozet
    return render_template('yillik_ozet.html', ozet=ozet)

@app.route('/export/yillik-pdf')
@sadece_yonetici
def export_yillik_pdf():
    rendered = render_template("yillik_ozet.html", ozet=session.get("ozet_cache", []))
    pdf = HTML(string=rendered).write_pdf()
    response = make_response(pdf)
    response.headers["Content-Disposition"] = "attachment; filename=yillik_ozet.pdf"
    response.headers["Content-type"] = "application/pdf"
    return response

@app.route("/raporlar")
@sadece_yonetici
def raporlar():
    if 'rol' not in session:
        return redirect('/')

    conn = sqlite3.connect("apartman.db")
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # Filtre parametreleri
    baslangic = request.args.get("baslangic")
    bitis = request.args.get("bitis")
    siralama = request.args.get("siralama", "tarih")
    yon = request.args.get("yon", "ASC")

    # GELİR sorgusu
    gelir_query = f"SELECT * FROM gelirler WHERE 1=1"
    gelir_params = []

    if baslangic:
        gelir_query += " AND tarih >= ?"
        gelir_params.append(baslangic)
    if bitis:
        gelir_query += " AND tarih <= ?"
        gelir_params.append(bitis)

    gelir_query += f" ORDER BY {siralama} {yon}"
    cur.execute(gelir_query, gelir_params)
    gelirler = cur.fetchall()

    # GİDER sorgusu
    gider_query = f"SELECT * FROM giderler WHERE 1=1"
    gider_params = []

    if baslangic:
        gider_query += " AND tarih >= ?"
        gider_params.append(baslangic)
    if bitis:
        gider_query += " AND tarih <= ?"
        gider_params.append(bitis)

    gider_query += f" ORDER BY {siralama} {yon}"
    cur.execute(gider_query, gider_params)
    giderler = cur.fetchall()

    # Aidat gelirleri (aidatlar tablosundan)
    cur.execute("""
    SELECT SUM(tutar) AS toplam_aidat
    FROM aidatlar
    WHERE tarih BETWEEN ? AND ?
    """, (baslangic, bitis))
    aidat_toplam = cur.fetchone()["toplam_aidat"] or 0

    

    # Diğer gelirler (gelirler tablosundan, Aidat başlığı hariç)
    cur.execute("""
        SELECT SUM(tutar) AS diger_gelir
        FROM gelirler
        WHERE baslik != 'Aidat'
        AND tarih BETWEEN ? AND ?
    """, (baslangic, bitis))
    diger_gelir_toplam = cur.fetchone()["diger_gelir"] or 0

    # Gider toplamı
    cur.execute("""
        SELECT SUM(tutar) AS toplam_gider
        FROM giderler
        WHERE tarih BETWEEN ? AND ?
    """, (baslangic, bitis))
    gider_toplam = cur.fetchone()["toplam_gider"] or 0

    # Kasa hesapları
    gelirler_toplam = aidat_toplam + diger_gelir_toplam
    donem_kasa = gelirler_toplam - gider_toplam

    conn.close()

    return render_template("raporlar.html",
                           gelirler=gelirler,
                           giderler=giderler,
                           baslangic=baslangic,
                           bitis=bitis,
                           siralama=siralama,
                           yon=yon,
                           aidat_toplam=aidat_toplam,
                           diger_gelir_toplam=diger_gelir_toplam,
                           gelirler_toplam=gelirler_toplam,
                           gider_toplam=gider_toplam,
                           donem_kasa=donem_kasa)




@app.route("/export/gelirler/pdf")
@sadece_yonetici
def export_gelirler_pdf():
    # Filtre ve sıralama parametreleri
    baslangic = request.args.get("baslangic")
    bitis = request.args.get("bitis")
    siralama = request.args.get("siralama", "tarih")
    yon = request.args.get("yon", "DESC")

    # Geçerli sıralama alanları
    allowed_fields = ["tarih", "tutar", "baslik"]
    if siralama not in allowed_fields:
        siralama = "tarih"
    if yon not in ["ASC", "DESC"]:
        yon = "DESC"

    conn = sqlite3.connect("apartman.db")
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # Sorgu + filtreler
    query = "SELECT baslik, tarih, tutar, aciklama FROM gelirler WHERE 1=1"
    params = []

    if baslangic:
        query += " AND tarih >= ?"
        params.append(baslangic)
    if bitis:
        query += " AND tarih <= ?"
        params.append(bitis)

    query += f" ORDER BY {siralama} {yon}"
    cur.execute(query, params)
    gelirler = cur.fetchall()
    conn.close()

    # Toplam hesapla
    toplam = sum([row["tutar"] for row in gelirler]) if gelirler else 0
    def tl_format(m): return "{:,}".format(int(m)).replace(",", ".")

    # PDF şablonu renderla
    rendered = render_template(
        "rapor_pdf_gelirler.html",
        gelirler=gelirler,
        toplam=tl_format(toplam),
        tl_format=tl_format
    )

    # PDF döndür
    pdf = HTML(string=rendered).write_pdf()
    response = make_response(pdf)
    response.headers["Content-Disposition"] = "attachment; filename=gelir_raporu.pdf"
    response.headers["Content-Type"] = "application/pdf"
    return response


@app.route("/export/giderler/pdf")
@sadece_yonetici
def export_giderler_pdf():
    baslangic = request.args.get("baslangic")
    bitis = request.args.get("bitis")
    siralama = request.args.get("siralama", "tarih")
    yon = request.args.get("yon", "DESC")

    allowed_fields = ["tarih", "tutar", "kategori"]
    if siralama not in allowed_fields:
        siralama = "tarih"
    if yon not in ["ASC", "DESC"]:
        yon = "DESC"

    conn = sqlite3.connect("apartman.db")
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    query = "SELECT tarih, kategori, tutar, aciklama FROM giderler WHERE 1=1"
    params = []
    if baslangic:
        query += " AND tarih >= ?"
        params.append(baslangic)
    if bitis:
        query += " AND tarih <= ?"
        params.append(bitis)
    query += f" ORDER BY {siralama} {yon}"

    cur.execute(query, params)
    giderler = cur.fetchall()
    conn.close()

    toplam = sum([row["tutar"] for row in giderler]) if giderler else 0
    def tl_format(miktar): return "{:,}".format(int(miktar)).replace(",", ".")

    rendered = render_template("rapor_pdf_giderler.html",
                               giderler=giderler,
                               toplam=tl_format(toplam),
                               tl_format=tl_format)

    pdf = HTML(string=rendered).write_pdf()
    response = make_response(pdf)
    response.headers["Content-Disposition"] = "attachment; filename=gider_raporu.pdf"
    response.headers["Content-Type"] = "application/pdf"
    return response

@app.route("/export/giderler")
@sadece_yonetici
def export_giderler():
    baslangic = request.args.get("baslangic")
    bitis = request.args.get("bitis")
    siralama = request.args.get("siralama", "tarih")
    yon = request.args.get("yon", "DESC")

    allowed_fields = ["tarih", "tutar", "kategori"]
    if siralama not in allowed_fields:
        siralama = "tarih"
    if yon not in ["ASC", "DESC"]:
        yon = "DESC"

    conn = sqlite3.connect("apartman.db")
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    query = "SELECT tarih, kategori, tutar, aciklama FROM giderler WHERE 1=1"
    params = []

    if baslangic:
        query += " AND tarih >= ?"
        params.append(baslangic)
    if bitis:
        query += " AND tarih <= ?"
        params.append(bitis)

    query += f" ORDER BY {siralama} {yon}"
    cur.execute(query, params)
    giderler = cur.fetchall()
    conn.close()

    from openpyxl import Workbook
    from io import BytesIO
    wb = Workbook()
    ws = wb.active
    ws.title = "Gider Raporu"

    ws.append(["Tarih", "Kategori", "Tutar", "Açıklama"])
    for row in giderler:
        ws.append([row["tarih"], row["kategori"], row["tutar"], row["aciklama"]])

    toplam = sum([r["tutar"] for r in giderler]) if giderler else 0
    ws.append(["", "", toplam, "Toplam Gider"])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    response = make_response(stream.read())
    response.headers["Content-Disposition"] = "attachment; filename=gider_raporu.xlsx"
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response



@app.route("/aidat-sil/<int:aidat_id>")
@sadece_yonetici
def aidat_sil(aidat_id):
    conn = sqlite3.connect("apartman.db")
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # aidatlar tablosunda primary key muhtemelen 'id'
    cur.execute("SELECT durum FROM aidatlar WHERE aidat_id = ?", (aidat_id,))
    row = cur.fetchone()

    if row and row["durum"] == "odendi":
        # gelirler tablosunda aidat_id varsa sil (opsiyonel)
        cur.execute("DELETE FROM gelirler WHERE aidat_id = ?", (aidat_id,))

    # aidatlar kaydını sil
    cur.execute("DELETE FROM aidatlar WHERE aidat_id = ?", (aidat_id,))
    conn.commit()
    conn.close()

    flash("Aidat başarıyla silindi.")
    return redirect("/aidatlar")




@app.route("/aidat-duzenle/<int:aidat_id>", methods=["GET", "POST"])
@sadece_yonetici
def aidat_duzenle(aidat_id):
    conn = sqlite3.connect("apartman.db")
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    if request.method == "POST":
        tarih = request.form.get("tarih")
        tutar = request.form.get("tutar")
        durum = request.form.get("durum")
        aciklama = request.form.get("aciklama")

        if durum not in ["odendi", "beklemede", "kismi"]:
            flash("Geçersiz durum.")
            return redirect(f"/aidat-duzenle/{aidat_id}")

        from datetime import datetime
        try:
            dt = datetime.strptime(tarih, "%Y-%m-%d")
            yil, ay = dt.year, dt.month
        except:
            flash("Geçersiz tarih formatı.")
            return redirect(f"/aidat-duzenle/{aidat_id}")

        cur.execute("""
            UPDATE aidatlar
            SET tarih = ?, tutar = ?, durum = ?, aciklama = ?, yil = ?, ay = ?
            WHERE aidat_id = ?
        """, (tarih, tutar, durum, aciklama, yil, ay, aidat_id))

        conn.commit()
        conn.close()
        flash("Aidat güncellendi.")
        return redirect("/aidatlar")

    # GET: Formu doldur
    cur.execute("""
        SELECT a.*, u.ad_soyad, u.daire_no
        FROM aidatlar a
        JOIN users u ON a.user_id = u.user_id
        WHERE a.aidat_id = ?
    """, (aidat_id,))
    kayit = cur.fetchone()
    conn.close()

    if not kayit:
        flash("Aidat kaydı bulunamadı.")
        return redirect("/aidatlar")

    return render_template("aidat_duzenle.html", kayit=kayit)



@app.route("/gelirler")

@sadece_yonetici
def gelir_listele():
    conn = sqlite3.connect("apartman.db")
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    cur.execute("""
        SELECT gelir_id, baslik, tutar, tarih, aciklama
        FROM gelirler
        ORDER BY tarih DESC
    """)
    gelirler = cur.fetchall()
    conn.close()

    return render_template("gelir_listele.html", gelirler=gelirler)



@app.route("/gelir-sil/<int:gelir_id>")
@sadece_yonetici
def gelir_sil(gelir_id):
    conn = sqlite3.connect("apartman.db")
    cur = conn.cursor()
    cur.execute("DELETE FROM gelirler WHERE gelir_id = ?", (gelir_id,))
    conn.commit()
    conn.close()
    flash("Gelir kaydı silindi.")
    return redirect("/gelirler")



@app.route("/gelir-duzenle/<int:gelir_id>", methods=["GET", "POST"])
@sadece_yonetici
def gelir_duzenle(gelir_id):
    conn = sqlite3.connect("apartman.db")
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    if request.method == "POST":
        baslik = request.form.get("baslik")
        tutar = request.form.get("tutar")
        tarih = request.form.get("tarih")
        aciklama = request.form.get("aciklama")

        cur.execute("""
            UPDATE gelirler
            SET baslik = ?, tutar = ?, tarih = ?, aciklama = ?
            WHERE gelir_id = ?
        """, (baslik, tutar, tarih, aciklama, gelir_id))

        conn.commit()
        conn.close()
        flash("Gelir kaydı güncellendi.")
        return redirect("/gelirler")

    # GET → Formu doldurmak için gelir bilgisi
    cur.execute("SELECT * FROM gelirler WHERE gelir_id = ?", (gelir_id,))
    kayit = cur.fetchone()
    conn.close()

    if not kayit:
        flash("Gelir kaydı bulunamadı.")
        return redirect("/gelirler")

    return render_template("gelir_duzenle.html", kayit=kayit)


@app.route("/giderler")
@sadece_yonetici
def gider_listele():
    conn = sqlite3.connect("apartman.db")
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    cur.execute("""
        SELECT id, kategori, tutar, tarih, aciklama
        FROM giderler
        ORDER BY tarih DESC
    """)
    kayitlar = cur.fetchall()
    conn.close()

    return render_template("gider_listele.html", giderler=kayitlar)


@app.route("/gider-sil/<int:id>")
@sadece_yonetici
def gider_sil(id):
    conn = sqlite3.connect("apartman.db")
    cur = conn.cursor()
    cur.execute("DELETE FROM giderler WHERE id = ?", (id,))
    conn.commit()
    conn.close()
    flash("Gider kaydı silindi.")
    return redirect("/giderler")


@app.route("/gider-duzenle/<int:id>", methods=["GET", "POST"])
@sadece_yonetici
def gider_duzenle(id):
    conn = sqlite3.connect("apartman.db")
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    if request.method == "POST":
        kategori = request.form.get("kategori")
        tutar = request.form.get("tutar")
        tarih = request.form.get("tarih")
        aciklama = request.form.get("aciklama")

        cur.execute("""
            UPDATE giderler
            SET kategori = ?, tutar = ?, tarih = ?, aciklama = ?
            WHERE id = ?
        """, (kategori, tutar, tarih, aciklama, id))

        conn.commit()
        conn.close()
        flash("Gider kaydı güncellendi.")
        return redirect("/giderler")

    cur.execute("SELECT * FROM giderler WHERE id = ?", (id,))
    kayit = cur.fetchone()
    conn.close()

    if not kayit:
        flash("Gider kaydı bulunamadı.")
        return redirect("/giderler")

    return render_template("gider_duzenle.html", kayit=kayit)




@app.route("/kisi-duzenle/<int:user_id>", methods=["GET", "POST"])
@sadece_yonetici
def kisi_duzenle(user_id):
    conn = sqlite3.connect("apartman.db")
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    if request.method == "POST":
        ad_soyad = request.form.get("ad_soyad")
        daire_no = request.form.get("daire_no")
        telefon = request.form.get("telefon")
        rol = request.form.get("rol")
        sifre = request.form.get("sifre")

        # Eğer şifre girildiyse güncelle, yoksa eski şifreyi koru
        if sifre:
            cur.execute("""
                UPDATE users
                SET ad_soyad = ?, daire_no = ?, telefon = ?, rol = ?, sifre = ?
                WHERE user_id = ?
            """, (ad_soyad, daire_no, telefon, rol, sifre, user_id))
        else:
            cur.execute("""
                UPDATE users
                SET ad_soyad = ?, daire_no = ?, telefon = ?, rol = ?
                WHERE user_id = ?
            """, (ad_soyad, daire_no, telefon, rol, user_id))

        conn.commit()
        conn.close()
        flash("Kişi bilgileri güncellendi.")
        return redirect("/kisiler")

    # GET metodu → formu doldurmak için veri çek
    cur.execute("SELECT * FROM users WHERE user_id = ?", (user_id,))
    kayit = cur.fetchone()
    conn.close()

    if not kayit:
        flash("Kişi bulunamadı.")
        return redirect("/kisiler")

    return render_template("kisi_duzenle.html", kayit=kayit)


@app.route("/kisiler")
@sadece_yonetici
@sadece_yonetici
def kisi_listele():
    conn = sqlite3.connect("apartman.db")
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT user_id, ad_soyad, daire_no, telefon, rol FROM users ORDER BY ad_soyad ASC")
    kisiler = cur.fetchall()
    conn.close()
    return render_template("kisi_listele.html", kisiler=kisiler)



@app.route("/kisi-sil/<int:user_id>")
@sadece_yonetici
def kisi_sil(user_id):
    conn = sqlite3.connect("apartman.db")
    cur = conn.cursor()
    cur.execute("DELETE FROM users WHERE user_id = ?", (user_id,))
    conn.commit()
    conn.close()
    flash("Kişi başarıyla silindi.")
    return redirect("/kisiler")



@app.route("/kisi-ekle", methods=["GET", "POST"])
@sadece_yonetici
def kisi_ekle():
    if request.method == "POST":
        ad_soyad = request.form.get("ad_soyad")
        daire_no = request.form.get("daire_no")
        telefon = request.form.get("telefon")
        sifre = request.form.get("sifre")
        rol = request.form.get("rol")

        conn = sqlite3.connect("apartman.db")
        cur = conn.cursor()

        cur.execute("""
            INSERT INTO users (ad_soyad, daire_no, telefon, sifre, rol)
            VALUES (?, ?, ?, ?, ?)
        """, (ad_soyad, daire_no, telefon, sifre, rol))

        conn.commit()
        conn.close()
        flash("Yeni kişi eklendi.")
        return redirect("/kisiler")

    return render_template("kisi_ekle.html")


@app.route("/ayarlar", methods=["GET", "POST"])
@sadece_yonetici
def ayarlar():
    conn = sqlite3.connect("apartman.db")
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    if request.method == "POST":
        aidat_ucreti = request.form.get("aidat_ucreti")
        mail_adresi = request.form.get("mail_adresi")

        cur.execute("UPDATE ayarlar SET aidat_ucreti = ?, mail_adresi = ? WHERE id = 1",
                    (aidat_ucreti, mail_adresi))
        conn.commit()
        flash("Ayarlar güncellendi.")

    cur.execute("SELECT * FROM ayarlar WHERE id = 1")
    kayit = cur.fetchone()
    conn.close()

    return render_template("ayarlar.html", ayar=kayit)



@app.route("/tarife-ekle", methods=["GET", "POST"])
@sadece_yonetici
def tarife_ekle():
    if request.method == "POST":
        from datetime import datetime
        import sqlite3

        baslangic = request.form.get("baslangic")  # '2025-01'
        bitis     = request.form.get("bitis")      # '2025-07'
        tutar     = float(request.form.get("tutar"))

        b_yil, b_ay = map(int, baslangic.split('-'))
        t_yil, t_ay = map(int, bitis.split('-'))

        conn = sqlite3.connect("apartman.db")
        cur = conn.cursor()

        yil, ay = b_yil, b_ay
        while (yil < t_yil) or (yil == t_yil and ay <= t_ay):
            # 📌 Bu yıl–ay için tarife var mı?
            cur.execute("SELECT COUNT(*) FROM aidat_tarifesi WHERE yil=? AND ay=?", (yil, ay))
            var_mi = cur.fetchone()[0]

            if var_mi:
                # 🔁 Varsa güncelle
                cur.execute("UPDATE aidat_tarifesi SET tutar=? WHERE yil=? AND ay=?", (tutar, yil, ay))
            else:
                # 🆕 Yoksa ekle
                cur.execute("INSERT INTO aidat_tarifesi (yil, ay, tutar) VALUES (?, ?, ?)", (yil, ay, tutar))

            ay += 1
            if ay > 12:
                ay = 1
                yil += 1

        conn.commit()
        conn.close()
        flash("Tarife başarıyla eklendi/güncellendi.")
        return redirect("/tarifeler")

    return render_template("aidat_tarife_ekle.html")



@app.route("/tarifeler")
@sadece_yonetici
def tarife_listele():
    import sqlite3
    conn = sqlite3.connect("apartman.db")
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT * FROM aidat_tarifesi ORDER BY yil DESC, ay DESC")

    tarifeler = cur.fetchall()
    conn.close()
    return render_template("tarife_listele.html", tarifeler=tarifeler)





@app.route("/tarife-duzenle/<int:tarife_id>", methods=["GET", "POST"])
@sadece_yonetici
def tarife_duzenle(tarife_id):
    import sqlite3
    conn = sqlite3.connect("apartman.db")
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    if request.method == "POST":
        yil   = int(request.form.get("yil"))
        ay    = int(request.form.get("ay"))
        tutar = float(request.form.get("tutar"))

        cur.execute("""
            UPDATE aidat_tarifesi
            SET yil = ?, ay = ?, tutar = ?
            WHERE id = ?
        """, (yil, ay, tutar, tarife_id))

        conn.commit()
        conn.close()
        flash("Tarife güncellendi.")
        return redirect("/tarifeler")

    # GET işlemi – mevcut veriyi getir
    cur.execute("SELECT * FROM aidat_tarifesi WHERE id = ?", (tarife_id,))
    tarife = cur.fetchone()
    conn.close()
    return render_template("tarife_duzenle.html", tarife=tarife)




@app.route("/tarife-sil/<int:tarife_id>")
@sadece_yonetici
def tarife_sil(tarife_id):
    import sqlite3
    conn = sqlite3.connect("apartman.db")
    cur = conn.cursor()

    cur.execute("DELETE FROM aidat_tarifesi WHERE id = ?", (tarife_id,))
    conn.commit()
    conn.close()
    flash("Tarife silindi.")
    return redirect("/tarifeler")



"""
@app.route("/borc-analiz", methods=["GET", "POST"])
@sadece_yonetici
def borc_analiz_sec():
    import sqlite3
    from datetime import datetime

    conn = sqlite3.connect("apartman.db")
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # 📋 USERS tablosundan kullanıcıları al
    cur.execute("SELECT user_id, ad_soyad, daire_no FROM users ORDER BY ad_soyad")
    kisiler = cur.fetchall()

    borclar = []
    toplam_borc = 0
    secilen_id = None

    if request.method == "POST":
        secilen_id = int(request.form.get("user_id"))
        now = datetime.now()

        cur.execute("
            SELECT t.yil, t.ay, t.tutar AS borc,
                   COALESCE((
                       SELECT SUM(a.tutar)
                       FROM aidatlar a
                       WHERE a.user_id = ? AND a.yil = t.yil AND a.ay = t.ay
                   ), 0) AS odenen
            FROM aidat_tarifesi t
            ORDER BY t.yil, t.ay
        ", (secilen_id,))
        borclar = cur.fetchall()
        toplam_borc = sum(b["borc"] for b in borclar)
        toplam_odenen = sum(b["odenen"] for b in borclar)


        toplam_borc = sum(b["tutar"] for b in borclar)

    conn.close()
    return render_template("borc_analiz_sec.html", kisiler=kisiler,
                           borclar=borclar, toplam_borc=toplam_borc,
                           secilen_id=secilen_id)

"""

@app.route("/toplu-odeme", methods=["GET", "POST"])
@sadece_yonetici
def toplu_odeme():
    import sqlite3, urllib.parse
    from datetime import datetime

    now = datetime.now()
    yil, ay = now.year, now.month
    tarih = now.strftime("%Y-%m-%d %H:%M:%S")

    conn = sqlite3.connect("apartman.db")
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    cur.execute("SELECT user_id, ad_soyad, daire_no, telefon FROM users ORDER BY ad_soyad")
    kisiler = cur.fetchall()

    secilen_id = None
    borclar = []
    detaylar = []
    toplam_borc = toplam_odenen = toplam_odeme = kalan_tutar = 0
    whatsapp_link = ""
    kisi = None

    if request.method == "POST":
        try:
            secilen_id = int(request.form.get("user_id"))
            odeme_raw = request.form.get("odeme")
            odeme = float(odeme_raw) if odeme_raw else 0
        except (TypeError, ValueError):
            odeme = 0

        kalan_tutar = toplam_odeme = odeme

        # 🔁 Kısmi kayıt varsa önce güncelle
        cur.execute("""
            SELECT id, yil, ay, tutar FROM aidatlar
            WHERE user_id = ? AND durum = 'kismi'
            ORDER BY yil, ay LIMIT 1
        """, (secilen_id,))
        kismi = cur.fetchone()
        if kismi and kalan_tutar > 0:
            y, a, mevcut, mevcut_id = kismi["yil"], kismi["ay"], kismi["tutar"], kismi["id"]
            cur.execute("SELECT tutar FROM aidat_tarifesi WHERE yil = ? AND ay = ?", (y, a))
            aidat_miktari = cur.fetchone()[0]
            eksik = aidat_miktari - mevcut
            odenecek = min(kalan_tutar, eksik)
            yeni_toplam = mevcut + odenecek
            yeni_durum = "odendi" if yeni_toplam >= aidat_miktari else "kismi"
            aciklama = f"{mevcut:.2f}+{odenecek:.2f}"
            cur.execute("""
                UPDATE aidatlar SET tutar = ?, durum = ?, aciklama = ?
                WHERE id = ?
            """, (yeni_toplam, yeni_durum, aciklama, mevcut_id))
            kalan_tutar -= odenecek
            detaylar.append(f"{y}/{str(a).zfill(2)} → ₺{odenecek:.2f} → {yeni_durum}")

        # 📥 Bekleyen aylar → kayıt oluştur
        cur.execute("""
            SELECT t.yil, t.ay, t.tutar FROM aidat_tarifesi t
            WHERE NOT EXISTS (
                SELECT 1 FROM aidatlar
                WHERE user_id = ? AND yil = t.yil AND ay = t.ay
            )
            AND (t.yil < ? OR (t.yil = ? AND t.ay <= ?))
            ORDER BY t.yil, t.ay
        """, (secilen_id, yil, yil, ay))
        bekleyenler = cur.fetchall()

        for b in bekleyenler:
            if kalan_tutar <= 0:
                break
            y, a, tutar = b["yil"], b["ay"], b["tutar"]
            odenecek = min(kalan_tutar, tutar)
            kalan_tutar -= odenecek
            durum = "odendi" if odenecek >= tutar else "kismi"
            aciklama = f"{odenecek:.2f}"
            aidat_id = aidat_id_uret(secilen_id, y, a)
            cur.execute("""
                INSERT INTO aidatlar (aidat_id, user_id, tarih, tutar, durum, aciklama, yil, ay)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (aidat_id, secilen_id, tarih, odenecek, durum, aciklama, y, a))
            detaylar.append(f"{y}/{str(a).zfill(2)} → ₺{odenecek:.2f} → {durum}")

        # 🔁 Fazla ödeme → gelecek aylar
        cur.execute("""
            SELECT t.yil, t.ay, t.tutar FROM aidat_tarifesi t
            WHERE t.yil > ? OR (t.yil = ? AND t.ay > ?)
            ORDER BY t.yil, t.ay
        """, (yil, yil, ay))
        gelecek = cur.fetchall()

        for g in gelecek:
            if kalan_tutar <= 0:
                break
            y, a, tutar = g["yil"], g["ay"], g["tutar"]
            cur.execute("SELECT SUM(tutar) FROM aidatlar WHERE user_id = ? AND yil = ? AND ay = ?",
                        (secilen_id, y, a))
            mevcut_odeme = cur.fetchone()[0] or 0
            kalan = tutar - mevcut_odeme
            if kalan <= 0:
                continue
            odenecek = min(kalan_tutar, kalan)
            kalan_tutar -= odenecek
            durum = "odendi" if odenecek == kalan else "kismi"
            aciklama = f"Fazla → {odenecek:.2f}"
            aidat_id = aidat_id_uret(secilen_id, y, a)
            cur.execute("""
                INSERT INTO aidatlar (aidat_id, user_id, tarih, tutar, durum, aciklama, yil, ay)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (aidat_id, secilen_id, tarih, odenecek, durum, aciklama, y, a))
            detaylar.append(f"{y}/{str(a).zfill(2)} → ₺{odenecek:.2f} → {durum}")

        # 📊 Panel tablosu → sadece bugüne kadar
        cur.execute("""
            SELECT t.yil, t.ay, t.tutar AS aidat_miktari,
                   COALESCE((SELECT SUM(a.tutar) FROM aidatlar a
                             WHERE a.user_id = ? AND a.yil = t.yil AND a.ay = t.ay), 0) AS odenen
            FROM aidat_tarifesi t
            WHERE t.yil < ? OR (t.yil = ? AND t.ay <= ?)
            ORDER BY t.yil, t.ay
        """, (secilen_id, yil, yil, ay))
        
        borclar = [b for b in cur.fetchall() if b["aidat_miktari"] - b["odenen"] > 0]

        toplam_borc = sum(b["aidat_miktari"] for b in borclar)
        toplam_odenen = sum(b["odenen"] for b in borclar)
        kalan_borc = toplam_borc - toplam_odenen

        # 📲 WhatsApp mesajı
        cur.execute("SELECT ad_soyad, daire_no, telefon FROM users WHERE user_id = ?", (secilen_id,))
        kisi = cur.fetchone()

        borc_aylari = [b for b in borclar if b["aidat_miktari"] - b["odenen"] > 0]
        borc_detay = "\n".join([
            f"{b['yil']}/{str(b['ay']).zfill(2)} → ₺{b['aidat_miktari']:.2f} - ₺{b['odenen']:.2f} = ₺{b['aidat_miktari'] - b['odenen']:.2f} borç"
            for b in borc_aylari
        ])

        mesaj = f"""SAYIN {kisi['ad_soyad']}
        {kisi['daire_no']} NOLU DAİRE AİDAT ÖDEME BİLGİLERİ

        📅 ÖDEME TARİHİ: {tarih}

        📊 ÖDEME DAĞILIMI:
        """ + "\n".join(detaylar) + f"""

        💳 TOPLAM ÖDEME: ₺{toplam_odeme:.2f}
        📌 KALAN BORÇ: ₺{kalan_borc:.2f}

        📋 BORÇ DETAYI:
        {borc_detay}

        🙏 Teşekkür eder, sağlıklı günler dileriz."""
                

        msg_encoded = urllib.parse.quote(mesaj)
        whatsapp_link = f"https://wa.me/{kisi['telefon'].replace(' ', '').replace('-', '')}?text={msg_encoded}"
        conn.commit()

    conn.close()
    return render_template("toplu_odeme.html",
                           kisiler=kisiler,
                           secilen_id=secilen_id,
                           borclar=borclar,
                           toplam_borc=toplam_borc,
                           toplam_odenen=toplam_odenen,
                           toplam_odeme=toplam_odeme,
                           kalan_tutar=kalan_tutar,
                           detaylar=detaylar,
                           whatsapp_link=whatsapp_link,
                           kisi=kisi)
    

@app.route("/aidat12")
@sadece_yonetici
def aidat12_raporu():
    import sqlite3
    from datetime import date

    from flask import render_template, request

    conn = sqlite3.connect("apartman.db")
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # Bulunduğumuz yılı al
    bugun_yil = date.today().year

    # Eğer URL'de 'yil' parametresi yoksa bugünkü yılı kullan
    yil = request.args.get("yil", str(bugun_yil))

    aylar = [
        ("01", "Ocak"), ("02", "Şubat"), ("03", "Mart"), ("04", "Nisan"),
        ("05", "Mayıs"), ("06", "Haziran"), ("07", "Temmuz"), ("08", "Ağustos"),
        ("09", "Eylül"), ("10", "Ekim"), ("11", "Kasım"), ("12", "Aralık")
    ]
    ay_kodlari = [a[0] for a in aylar]

    cur.execute("SELECT user_id, daire_no, ad_soyad FROM users")
    sakinler = sorted(cur.fetchall(), key=lambda x: int(x["daire_no"]))

    tablo = {}
    for s in sakinler:
        ad = f"{int(s['daire_no']):02d} - {s['ad_soyad']}"
        tablo[ad] = {ay: 0 for ay in ay_kodlari}

    cur.execute("""
        SELECT u.daire_no, u.ad_soyad, a.ay, SUM(a.tutar) AS tutar
        FROM aidatlar a
        JOIN users u ON u.user_id = a.user_id
        WHERE a.yil = ?
        GROUP BY u.user_id, a.ay
    """, (yil,))
    veriler = cur.fetchall()

    conn.close()

    toplam_genel = 0
    for r in veriler:
        ad = f"{int(r['daire_no']):02d} - {r['ad_soyad']}"
        ay = str(r["ay"]).zfill(2)
        tutar = r["tutar"] or 0
        if ad in tablo:
            tablo[ad][ay] = tutar
            toplam_genel += tutar

    return render_template(
        "aidat12.html",
        tablo=tablo,
        aylar=aylar,
        yil=yil,
        toplam_genel=toplam_genel
    )


@app.route("/aidat12/excel")
def aidat12_excel():
    import sqlite3
    import pandas as pd
    from flask import request, send_file
    from io import BytesIO
    from datetime import date


    # Bulunduğumuz yılı al
    bugun_yil = date.today().year

    # Eğer URL'de 'yil' parametresi yoksa bugünkü yılı kullan
    yil = request.args.get("yil", str(bugun_yil))

    conn = sqlite3.connect("apartman.db")
    cur = conn.cursor()

    cur.execute("""
        SELECT u.daire_no, u.ad_soyad, a.ay, SUM(a.tutar)
        FROM aidatlar a
        JOIN users u ON u.user_id = a.user_id
        WHERE a.yil = ?
        GROUP BY u.user_id, a.ay
    """, (yil,))
    rows = cur.fetchall()
    conn.close()

    data = {}
    for daire, ad, ay, tutar in rows:
        key = f"{int(daire):02d} - {ad}"
        ay = str(ay).zfill(2)
        if key not in data:
            data[key] = {}
        data[key][ay] = tutar

    aylar = ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12"]
    df = pd.DataFrame([
        {"Sakin": k, **{a: data[k].get(a, 0) for a in aylar}}
        for k in data
    ])

    output = BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False, sheet_name=f"Aidat_{yil}")

    output.seek(0)
    return send_file(output, as_attachment=True, download_name=f"aidat_{yil}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/aidat12/pdf")
def aidat12_pdf():
    import sqlite3
    from flask import request, send_file
    from io import BytesIO
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib import colors
    from datetime import date


    # 🔤 Türkçe karakter destekleyen fontu kaydet
    pdfmetrics.registerFont(TTFont("DejaVu", "DejaVuSans.ttf"))

    # 🟡 Seçilen yılı al
    # Bulunduğumuz yılı al
    bugun_yil = date.today().year

    # Eğer URL'de 'yil' parametresi yoksa bugünkü yılı kullan
    yil = request.args.get("yil", str(bugun_yil))

    # 💳 Verileri veritabanından çek
    conn = sqlite3.connect("apartman.db")
    cur = conn.cursor()

    cur.execute("""
        SELECT u.daire_no, u.ad_soyad, a.ay, SUM(a.tutar)
        FROM aidatlar a
        JOIN users u ON u.user_id = a.user_id
        WHERE a.yil = ?
        GROUP BY u.user_id, a.ay
    """, (yil,))
    rows = cur.fetchall()
    conn.close()

    # 📊 Tablo verisini hazırla
    data = {}
    for daire, ad, ay, tutar in rows:
        key = f"{int(daire):02d} - {ad}"
        ay = str(ay).zfill(2)
        if key not in data:
            data[key] = {}
        data[key][ay] = tutar

    aylar = ["01", "02", "03", "04", "05", "06",
             "07", "08", "09", "10", "11", "12"]
    header = ["Sakin"] + aylar
    table_rows = [[k] + [f"{data[k].get(a, 0):.2f}" for a in aylar] for k in data]
    table_data = [header] + table_rows

    # 📄 PDF hazırlığı
    pdf_buffer = BytesIO()
    doc = SimpleDocTemplate(
        pdf_buffer,
        pagesize=landscape(A4),
        rightMargin=20,
        leftMargin=20,
        topMargin=20,
        bottomMargin=20
    )

    # 📐 Sütun genişliklerini ayarla
    isim_genislik = 45 * mm
    ay_genislik = (A4[1] - 2 * 20 * mm - isim_genislik) / 12
    col_widths = [isim_genislik] + [ay_genislik] * 12

    # 🧾 Tabloyu oluştur
    table = Table(table_data, colWidths=col_widths)

    # 💄 Stil tanımı
    style = TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "DejaVu"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.black),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
    ])

    # 🎨 Hücrelere renklendirme uygulaması
    for row_index, row in enumerate(table_data[1:], start=1):  # Başlık satırı hariç
        for col_index, value in enumerate(row[1:], start=1):   # İlk sütun "Sakin" hariç
            try:
                tutar = float(value.replace(",", "").replace("₺", ""))
            except:
                tutar = 0
            renk = colors.HexColor("#F8D7DA") if tutar == 0 else colors.HexColor("#D1E7DD")
            style.add("BACKGROUND", (col_index, row_index), (col_index, row_index), renk)

    table.setStyle(style)

    # 📥 PDF dosyasını oluştur
    doc.build([table])
    pdf_buffer.seek(0)

    return send_file(
        pdf_buffer,
        as_attachment=True,
        download_name=f"aidat_{yil}.pdf",
        mimetype="application/pdf"
    )

@app.route("/veritabani/indir")
@sadece_yonetici
def indir_db():
    return send_file("apartman.db", as_attachment=True)


@app.route("/veritabani/yukle", methods=["POST"])
@sadece_yonetici
def yukle_db():
    from flask import request
    dosya = request.files["dbfile"]
    if dosya.filename.endswith(".db"):
        dosya.save("apartman.db")
        return "✅ Veritabanı başarıyla yüklendi"
    else:
        return "❌ Sadece .db dosyaları kabul edilir", 400




if __name__ == '__main__':
    app.run(debug=True)
